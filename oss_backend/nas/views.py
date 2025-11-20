from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io
from django.core.exceptions import ValidationError

from .filters import VendorFilter
from .models import Nas, NasGroup, Vendor
from .serializers import (
    NasSerializer, NasCreateSerializer, NasUpdateSerializer,
    NasGroupSerializer, NasGroupTreeSerializer, VendorSerializer
)


class NasGroupViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing NAS groups.
    """
    queryset = NasGroup.objects.all()
    serializer_class = NasGroupSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['name', 'parent']
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created_at']
    ordering = ['name']

    @action(detail=False, methods=['get'])
    def tree(self, request):
        """
        Return a tree structure of all NAS groups.
        """
        root_nodes = NasGroup.objects.root_nodes()
        serializer = NasGroupTreeSerializer(root_nodes, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def list_all(self, request):
        """
        Return a list of all vendors without pagination.
        """
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class NasViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing NAS devices.
    """
    queryset = Nas.objects.all()
    serializer_class = NasSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['name', 'ip_address', 'coa_enabled', 'is_active', 'groups']
    search_fields = ['name', 'description', 'ip_address']
    ordering_fields = ['name', 'created_at', 'ip_address']
    ordering = ['name']

    def get_serializer_class(self):
        """
        Return appropriate serializer class based on the request method.
        """
        if self.action == 'create':
            return NasCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return NasUpdateSerializer
        return NasSerializer

    @action(detail=False, methods=['get'])
    def by_group(self, request):
        """
        Filter NAS devices by group ID.
        """
        group_id = request.query_params.get('group_id')
        if not group_id:
            return Response(
                {"error": "group_id parameter is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            group = NasGroup.objects.get(pk=group_id)
            # Get all descendant groups including the current group
            group_ids = [group.id] + [g.id for g in group.get_descendants()]
            nas_devices = Nas.objects.filter(groups__id__in=group_ids).distinct()
            serializer = self.get_serializer(nas_devices, many=True)
            return Response(serializer.data)
        except NasGroup.DoesNotExist:
            return Response(
                {"error": f"NasGroup with id {group_id} does not exist"},
                status=status.HTTP_404_NOT_FOUND
            )

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """
        Download Excel template for importing NAS devices.
        """
        from shared.models import Timezone
        from radius.models import Secret

        wb = Workbook()
        ws = wb.active
        ws.title = "NAS Devices"

        # Header row
        headers = ['Name', 'IP Address', 'Description', 'Vendor ID', 'Secret ID', 
                   'Timezone ID', 'CoA Enabled', 'CoA Port', 'Group Names (comma-separated)', 'Is Active']
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Example row
        vendor_id = Vendor.objects.first().id if Vendor.objects.exists() else 1
        secret_id = Secret.objects.first().id if Secret.objects.exists() else None
        timezone_id = Timezone.objects.first().id if Timezone.objects.exists() else 1
        example_row = [
            'Example NAS',
            '192.168.1.100',
            'Example NAS device description',
            vendor_id,
            secret_id or '',
            timezone_id,
            'true',
            '3799',
            '',
            'true'
        ]
        for col, value in enumerate(example_row, start=1):
            ws.cell(row=2, column=col, value=value)

        # Adjust column widths
        column_widths = [25, 20, 40, 12, 12, 12, 15, 12, 35, 12]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Create response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="nas_devices_template.xlsx"'
        return response

    @action(detail=False, methods=['post'])
    def upload_excel(self, request):
        """
        Upload and import NAS devices from Excel file.
        """
        from shared.models import Timezone
        from radius.models import Secret
        from django.apps import apps
        import ipaddress
        import re

        if 'file' not in request.FILES:
            return Response(
                {"error": "No file provided"},
                status=status.HTTP_400_BAD_REQUEST
            )

        file = request.FILES['file']
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {"error": "Invalid file format. Please upload an Excel file (.xlsx or .xls)"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename=io.BytesIO(file.read()))
            ws = wb.active

            created = []
            errors = []

            # Skip header row (row 1) and process from row 2
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Skip empty rows
                    continue

                name = row[0] if len(row) > 0 else None
                ip_address = row[1] if len(row) > 1 else None
                description = row[2] if len(row) > 2 else ''
                vendor_id = row[3] if len(row) > 3 else None
                secret_id = row[4] if len(row) > 4 else None
                timezone_id = row[5] if len(row) > 5 else None
                coa_enabled = str(row[6]).lower() == 'true' if len(row) > 6 and row[6] else False
                coa_port = int(row[7]) if len(row) > 7 and row[7] else 3799
                group_names_str = row[8] if len(row) > 8 and row[8] else ''
                is_active = str(row[9]).lower() != 'false' if len(row) > 9 and row[9] else True

                # Validate required fields
                if not name:
                    errors.append(f"Row {row_num}: Name is required")
                    continue
                if not ip_address:
                    errors.append(f"Row {row_num}: IP Address is required")
                    continue
                if not vendor_id:
                    errors.append(f"Row {row_num}: Vendor ID is required")
                    continue
                if not timezone_id:
                    errors.append(f"Row {row_num}: Timezone ID is required")
                    continue

                # Validate IP address
                try:
                    ipaddress.ip_address(str(ip_address))
                except ValueError:
                    # Try to validate as hostname
                    hostname_pattern = re.compile(
                        r'^([a-zA-Z0-9]([a-zA-Z0-9\-]{0,61}[a-zA-Z0-9])?\.)*[a-zA-Z0-9]([a-zA-Z0-9\-]{0,61}[a-zA-Z0-9])?$'
                    )
                    simple_hostname_pattern = re.compile(r'^[a-zA-Z0-9]([a-zA-Z0-9\-_]{0,61}[a-zA-Z0-9])?$')
                    if not (hostname_pattern.match(str(ip_address)) or simple_hostname_pattern.match(str(ip_address))):
                        errors.append(f"Row {row_num}: Invalid IP address or hostname")
                        continue

                # Validate vendor exists
                try:
                    vendor = Vendor.objects.get(pk=vendor_id)
                except Vendor.DoesNotExist:
                    errors.append(f"Row {row_num}: Vendor with ID {vendor_id} not found")
                    continue

                # Validate timezone exists
                try:
                    timezone = Timezone.objects.get(pk=timezone_id)
                except Timezone.DoesNotExist:
                    errors.append(f"Row {row_num}: Timezone with ID {timezone_id} not found")
                    continue

                # Validate secret if provided
                secret = None
                if secret_id:
                    try:
                        secret = Secret.objects.get(pk=secret_id)
                    except Secret.DoesNotExist:
                        errors.append(f"Row {row_num}: Secret with ID {secret_id} not found")
                        continue

                # Get groups by names
                group_ids = []
                if group_names_str:
                    group_names = [g.strip() for g in str(group_names_str).split(',') if g.strip()]
                    for group_name in group_names:
                        try:
                            group = NasGroup.objects.get(name=group_name)
                            group_ids.append(group.id)
                        except NasGroup.DoesNotExist:
                            errors.append(f"Row {row_num}: Group '{group_name}' not found")
                            continue

                # Create NAS device
                try:
                    nas_data = {
                        'name': name,
                        'ip_address': str(ip_address),
                        'description': description or '',
                        'vendor_id': vendor_id,
                        'timezone_id': timezone_id,
                        'coa_enabled': coa_enabled,
                        'coa_port': coa_port,
                        'is_active': is_active,
                        'group_ids': group_ids
                    }
                    if secret_id:
                        nas_data['secret_id'] = secret_id

                    serializer = NasCreateSerializer(data=nas_data)
                    if serializer.is_valid():
                        nas = serializer.save()
                        created.append(nas.name)
                    else:
                        errors.append(f"Row {row_num}: {serializer.errors}")
                except Exception as e:
                    errors.append(f"Row {row_num}: Error creating NAS - {str(e)}")

            return Response({
                "success": True,
                "created": len(created),
                "created_devices": created,
                "errors": errors
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": f"Error processing file: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

class VendorsViewSet(viewsets.ModelViewSet):
    """
    ViewSet for viewing and editing AdminUser instances.
    """
    queryset = Vendor.objects.all()
    serializer_class = VendorSerializer
    filterset_class = VendorFilter

    @action(detail=False, methods=['get'])
    def list_all(self, request):
        """
        Return a list of all vendors without pagination.
        """
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def get_serializer_class(self):
        """
        Return the appropriate serializer class based on the action.
        """
        return self.serializer_class

    def create(self, request, *args, **kwargs):
        """
        Create a new vendor.
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def perform_create(self, serializer):
        """
        Save the user and return the instance.
        """
        return serializer.save()

    def update(self, request, *args, **kwargs):
        """
        Update an existing vendor.
        """
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        if getattr(instance, '_prefetched_objects_cache', None):
            # If 'prefetch_related' has been applied to a queryset, we need to
            # forcibly invalidate the prefetch cache on the instance.
            instance._prefetched_objects_cache = {}

        return Response(serializer.data)