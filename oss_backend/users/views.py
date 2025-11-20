from rest_framework import viewsets, permissions
from rest_framework.response import Response
from rest_framework import status
from rest_framework import permissions

from drf_yasg.utils import swagger_auto_schema

from nas.models import Nas
from nas.serializers import NasSerializer
from .models import User, UserGroup, UserIdentifierType, UserIdentifier, UserIdentifierNasAuthorization
from .serializers import (
    UserSerializer, UserCreateSerializer, UserUpdateSerializer, UserGroupSerializer,
    UserGroupTreeSerializer, UserIdentifierTypeSerializer, UserIdentifierSerializer,
    UserIdentifierNasAuthorizationSerializer, UserIdentifierNasAuthorizationCreateSerializer,
    UserIdentifierNasAuthorizationUpdateSerializer
)
from rest_framework.decorators import action
from rest_framework import serializers
from rest_framework.views import APIView
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io
from datetime import datetime

class UserGroupViewSet(viewsets.ModelViewSet):
    """
    ViewSet for viewing and editing User instances.
    """
    queryset = UserGroup.objects.all()
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'create':
            return UserGroupSerializer
        elif self.action in ['update', 'partial_update']:
            return UserGroupSerializer
        return UserGroupSerializer

    @swagger_auto_schema(
        operation_description="List all user groups",
        responses={200: UserGroupSerializer(many=True)}
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @action(detail=False, methods=['get'])
    def list_all(self, request):
        """
        Return a list of all admin users without pagination.
        """
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @swagger_auto_schema(
        operation_description="Create a new user group",
        request_body=UserGroupSerializer,
        responses={201: UserGroupSerializer()}
    )
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(
            UserGroupSerializer(instance=serializer.instance).data,
            status=status.HTTP_201_CREATED,
            headers=headers
        )

    @swagger_auto_schema(
        operation_description="Retrieve a user group by ID",
        responses={200: UserGroupSerializer()}
    )
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_description="Update a user group",
        request_body=UserGroupSerializer,
        responses={200: UserGroupSerializer()}
    )
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(UserGroupSerializer(instance=instance).data)

    @swagger_auto_schema(
        operation_description="Delete a user group",
        responses={204: "No content"}
    )
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=['get'])
    def tree(self, request):
        queryset = self.get_queryset()
        # Assuming you have a serializer that can handle tree structure
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """
        Download Excel template for importing user groups.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "User Groups"

        # Header row
        headers = ['Name', 'Description', 'Parent Group Name', 'Allow Any NAS']
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Example row
        example_row = ['Example Group', 'This is an example group description', '', 'false']
        for col, value in enumerate(example_row, start=1):
            ws.cell(row=2, column=col, value=value)

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20

        # Create response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="user_groups_template.xlsx"'
        return response

    @action(detail=False, methods=['post'])
    def upload_excel(self, request):
        """
        Upload and import user groups from Excel file.
        """
        if 'file' not in request.FILES:
            return Response(
                {"error": "No file provided"},
                status=status.HTTP_400_BAD_REQUEST
            )

        file = request.FILES['file']
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {"error": "Invalid file format. Please upload an Excel file (.xlsx or .xls)"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            wb = load_workbook(filename=io.BytesIO(file.read()))
            ws = wb.active

            created = []
            errors = []

            # Skip header row (row 1) and process from row 2
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Skip empty rows
                    continue

                name = row[0] if len(row) > 0 else None
                description = row[1] if len(row) > 1 else ''
                parent_name = row[2] if len(row) > 2 else None
                allow_any_nas = str(row[3]).lower() == 'true' if len(row) > 3 and row[3] else False

                if not name:
                    errors.append(f"Row {row_num}: Name is required")
                    continue

                # Get or create parent group
                parent = None
                if parent_name:
                    try:
                        parent = UserGroup.objects.get(name=parent_name)
                    except UserGroup.DoesNotExist:
                        errors.append(f"Row {row_num}: Parent group '{parent_name}' not found")
                        continue

                # Create group
                try:
                    group, created_flag = UserGroup.objects.get_or_create(
                        name=name,
                        defaults={
                            'description': description or '',
                            'parent': parent,
                            'allow_any_nas': allow_any_nas
                        }
                    )
                    if created_flag:
                        created.append(group.name)
                    else:
                        errors.append(f"Row {row_num}: Group '{name}' already exists")
                except Exception as e:
                    errors.append(f"Row {row_num}: Error creating group - {str(e)}")

            return Response({
                "success": True,
                "created": len(created),
                "created_groups": created,
                "errors": errors
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": f"Error processing file: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

class UserIdentifierTypeViewSet(viewsets.ModelViewSet):
    queryset = UserIdentifierType.objects.all()
    serializer_class = UserIdentifierTypeSerializer
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=False, methods=['get'])
    def list_all(self, request):
        """
        Return a list of all admin users without pagination.
        """
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

class UserIdentifierViewSet(viewsets.ModelViewSet):
    serializer_class = UserIdentifierSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user_id = self.kwargs.get('user_pk')
        return UserIdentifier.objects.filter(user_id=user_id)

    def perform_create(self, serializer):
        user_id = self.kwargs.get('user_pk')
        serializer.save(user_id=user_id)

class UserViewSet(viewsets.ModelViewSet):
    """
    ViewSet for viewing and editing User instances.
    """
    queryset = User.objects.all()
    permission_classes = [permissions.IsAuthenticated]
    
    def get_serializer_class(self):
        if self.action == 'create':
            return UserCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return UserUpdateSerializer
        return UserSerializer
    
    @swagger_auto_schema(
        operation_description="List all users",
        responses={200: UserSerializer(many=True)}
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)
    
    @swagger_auto_schema(
        operation_description="Create a new user",
        request_body=UserCreateSerializer,
        responses={201: UserSerializer()}
    )
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(
            UserSerializer(instance=serializer.instance).data,
            status=status.HTTP_201_CREATED,
            headers=headers
        )
    
    @swagger_auto_schema(
        operation_description="Retrieve a user by ID",
        responses={200: UserSerializer()}
    )
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)
    
    @swagger_auto_schema(
        operation_description="Update a user",
        request_body=UserUpdateSerializer,
        responses={200: UserSerializer()}
    )
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(UserSerializer(instance=instance).data)
    
    @swagger_auto_schema(
        operation_description="Delete a user",
        responses={204: "No content"}
    )
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['get', 'post'])
    def identifiers(self, request, pk=None):
        user = self.get_object()
        if request.method == 'GET':
            identifiers = UserIdentifier.objects.filter(user=user)
            serializer = UserIdentifierSerializer(identifiers, many=True)
            return Response(serializer.data)
        elif request.method == 'POST':
            serializer = UserIdentifierSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save(user=user)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['get', 'put', 'delete'], url_path='identifiers/(?P<identifier_id>[^/.]+)')
    def identifier(self, request, pk=None, identifier_id=None):
        user = self.get_object()
        try:
            identifier = UserIdentifier.objects.get(user=user, id=identifier_id)
        except UserIdentifier.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)

        if request.method == 'GET':
            serializer = UserIdentifierSerializer(identifier)
            return Response(serializer.data)
        elif request.method == 'PUT':
            serializer = UserIdentifierSerializer(identifier, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        elif request.method == 'DELETE':
            identifier.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)

    def _handle_identifiers(self, user, identifiers_data):
        """Handle updating user identifiers."""
        if not identifiers_data:
            return

        # Get existing identifiers
        existing_identifiers = {str(identifier.id): identifier for identifier in user.identifiers.all()}
        
        # Create a map of existing identifiers by type and value
        existing_by_type_value = {
            (str(identifier.identifier_type_id), identifier.value): identifier 
            for identifier in user.identifiers.all()
        }
        
        # Process each identifier in the request
        for identifier_data in identifiers_data:
            identifier_id = identifier_data.get('id')
            type_id = str(identifier_data.get('identifier_type_id'))
            value = identifier_data.get('value')
            
            # Try to find existing identifier by ID or by type+value combination
            if identifier_id and str(identifier_id) in existing_identifiers:
                identifier = existing_identifiers[str(identifier_id)]
            elif (type_id, value) in existing_by_type_value:
                identifier = existing_by_type_value[(type_id, value)]
            else:
                identifier = None
            
            if identifier:
                # Update existing identifier
                serializer = UserIdentifierSerializer(identifier, data=identifier_data, partial=True)
                if serializer.is_valid():
                    serializer.save()
                else:
                    raise serializers.ValidationError(serializer.errors)
            else:
                # Create new identifier
                serializer = UserIdentifierSerializer(data=identifier_data)
                if serializer.is_valid():
                    serializer.save(user=user)
                else:
                    raise serializers.ValidationError(serializer.errors)

        # Remove identifiers that are no longer in the request
        current_ids = {str(data.get('id')) for data in identifiers_data if data.get('id')}
        current_type_values = {
            (str(data.get('identifier_type_id')), data.get('value'))
            for data in identifiers_data
        }
        
        for identifier_id, identifier in existing_identifiers.items():
            if (identifier_id not in current_ids and 
                (str(identifier.identifier_type_id), identifier.value) not in current_type_values):
                identifier.delete()

    def perform_update(self, serializer):
        """Handle the update of a user instance."""
        instance = serializer.instance
        data = serializer.validated_data

        # Handle groups update
        if 'groups' in data:
            # Clear existing groups and set new ones
            instance.groups.clear()
            instance.groups.add(*data['groups'])

        # Save the user instance
        instance = serializer.save()

        # Handle identifiers update
        identifiers_data = self.request.data.get('identifiers', [])
        self._handle_identifiers(instance, identifiers_data)

        return instance

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """
        Download Excel template for importing users with identifiers and NAS authorizations.
        Creates a workbook with three sheets: Users, Identifiers, NAS Authorizations
        """
        from radius.models import AuthAttributeGroup
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Sheet 1: Users
        ws_users = wb.create_sheet("Users")
        headers_users = [
            'Email', 'External ID', 'First Name', 'Last Name', 'Phone Number',
            'Group Names (comma-separated)', 'Allow Any NAS', 'Is Active'
        ]
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers_users, start=1):
            cell = ws_users.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Example row for users
        example_row_users = [
            'user@example.com', 'EXT001', 'John', 'Doe', '+1234567890',
            'Example Group', '', 'true'
        ]
        for col, value in enumerate(example_row_users, start=1):
            ws_users.cell(row=2, column=col, value=value)
        
        # Adjust column widths for users sheet
        column_widths_users = [30, 15, 20, 20, 20, 35, 15, 12]
        for col, width in enumerate(column_widths_users, start=1):
            ws_users.column_dimensions[get_column_letter(col)].width = width
        
        # Sheet 2: Identifiers
        ws_identifiers = wb.create_sheet("Identifiers")
        headers_identifiers = [
            'User Email or External ID', 'Identifier Type Code', 'Value', 'Password',
            'Is Enabled', 'Comment', 'Auth Attribute Group Name', 'Expiration Date',
            'Reject Expired', 'Expired Auth Attribute Group Name'
        ]
        
        for col, header in enumerate(headers_identifiers, start=1):
            cell = ws_identifiers.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Example row for identifiers
        example_row_identifiers = [
            'user@example.com', 'PWD', 'username', 'password123',
            'true', 'Example identifier', '', '', 'false', ''
        ]
        for col, value in enumerate(example_row_identifiers, start=1):
            ws_identifiers.cell(row=2, column=col, value=value)
        
        # Adjust column widths for identifiers sheet
        column_widths_identifiers = [30, 20, 25, 20, 12, 30, 30, 20, 15, 35]
        for col, width in enumerate(column_widths_identifiers, start=1):
            ws_identifiers.column_dimensions[get_column_letter(col)].width = width
        
        # Sheet 3: NAS Authorizations
        ws_auth = wb.create_sheet("NAS Authorizations")
        headers_auth = [
            'User Email or External ID', 'Identifier Value', 'NAS Name',
            'Auth Attribute Group Name'
        ]
        
        for col, header in enumerate(headers_auth, start=1):
            cell = ws_auth.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Example row for NAS authorizations
        example_row_auth = [
            'user@example.com', 'username', 'Example NAS', ''
        ]
        for col, value in enumerate(example_row_auth, start=1):
            ws_auth.cell(row=2, column=col, value=value)
        
        # Adjust column widths for NAS authorizations sheet
        column_widths_auth = [30, 25, 30, 30]
        for col, width in enumerate(column_widths_auth, start=1):
            ws_auth.column_dimensions[get_column_letter(col)].width = width
        
        # Create response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="users_template.xlsx"'
        return response

    @action(detail=False, methods=['post'])
    def upload_excel(self, request):
        """
        Upload and import users with identifiers and NAS authorizations from Excel file.
        """
        from nas.models import Nas
        from radius.models import AuthAttributeGroup
        from django.utils import timezone
        
        if 'file' not in request.FILES:
            return Response(
                {"error": "No file provided"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES['file']
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {"error": "Invalid file format. Please upload an Excel file (.xlsx or .xls)"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            wb = load_workbook(filename=io.BytesIO(file.read()))
            
            created_users = []
            errors = []
            user_map = {}  # Maps email/external_id to user objects
            
            # Process Users sheet
            if "Users" in wb.sheetnames:
                ws_users = wb["Users"]
                for row_num, row in enumerate(ws_users.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row):  # Skip empty rows
                        continue
                    
                    email = row[0] if len(row) > 0 else None
                    external_id = row[1] if len(row) > 1 else None
                    first_name = row[2] if len(row) > 2 else ''
                    last_name = row[3] if len(row) > 3 else ''
                    phone_number = row[4] if len(row) > 4 else ''
                    group_names_str = row[5] if len(row) > 5 else ''
                    allow_any_nas = None
                    if len(row) > 6 and row[6]:
                        allow_any_nas = str(row[6]).lower() == 'true'
                    is_active = str(row[7]).lower() != 'false' if len(row) > 7 and row[7] else True
                    
                    if not email:
                        errors.append(f"Users Sheet Row {row_num}: Email is required")
                        continue
                    
                    # Get groups
                    group_ids = []
                    if group_names_str:
                        group_names = [g.strip() for g in str(group_names_str).split(',') if g.strip()]
                        for group_name in group_names:
                            try:
                                group = UserGroup.objects.get(name=group_name)
                                group_ids.append(group.id)
                            except UserGroup.DoesNotExist:
                                errors.append(f"Users Sheet Row {row_num}: Group '{group_name}' not found")
                                continue
                    
                    # Create user
                    try:
                        user_data = {
                            'email': email,
                            'first_name': first_name or '',
                            'last_name': last_name or '',
                            'phone_number': phone_number or '',
                            'is_active': is_active,
                            'group_ids': group_ids
                        }
                        if external_id:
                            user_data['external_id'] = external_id
                        if allow_any_nas is not None:
                            user_data['allow_any_nas'] = allow_any_nas
                        
                        serializer = UserCreateSerializer(data=user_data)
                        if serializer.is_valid():
                            user = serializer.save()
                            created_users.append(user.email)
                            user_map[email] = user
                            if external_id:
                                user_map[external_id] = user
                        else:
                            errors.append(f"Users Sheet Row {row_num}: {serializer.errors}")
                    except Exception as e:
                        errors.append(f"Users Sheet Row {row_num}: Error creating user - {str(e)}")
            
            # Process Identifiers sheet
            if "Identifiers" in wb.sheetnames:
                ws_identifiers = wb["Identifiers"]
                for row_num, row in enumerate(ws_identifiers.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row):  # Skip empty rows
                        continue
                    
                    user_key = row[0] if len(row) > 0 else None  # Email or External ID
                    identifier_type_code = row[1] if len(row) > 1 else None
                    value = row[2] if len(row) > 2 else None
                    password = row[3] if len(row) > 3 else ''
                    is_enabled = str(row[4]).lower() != 'false' if len(row) > 4 and row[4] else True
                    comment = row[5] if len(row) > 5 else ''
                    auth_group_name = row[6] if len(row) > 6 else None
                    expiration_date_str = row[7] if len(row) > 7 else None
                    reject_expired = str(row[8]).lower() == 'true' if len(row) > 8 and row[8] else False
                    expired_auth_group_name = row[9] if len(row) > 9 else None
                    
                    if not user_key or not identifier_type_code or not value:
                        errors.append(f"Identifiers Sheet Row {row_num}: User key, Identifier Type Code, and Value are required")
                        continue
                    
                    # Find user
                    user = user_map.get(user_key)
                    if not user:
                        try:
                            user = User.objects.get(email=user_key)
                        except User.DoesNotExist:
                            try:
                                user = User.objects.get(external_id=user_key)
                            except User.DoesNotExist:
                                errors.append(f"Identifiers Sheet Row {row_num}: User with key '{user_key}' not found")
                                continue
                    
                    # Get identifier type
                    try:
                        identifier_type = UserIdentifierType.objects.get(code=identifier_type_code)
                    except UserIdentifierType.DoesNotExist:
                        errors.append(f"Identifiers Sheet Row {row_num}: Identifier type with code '{identifier_type_code}' not found")
                        continue
                    
                    # Get auth attribute groups
                    auth_group = None
                    if auth_group_name:
                        try:
                            auth_group = AuthAttributeGroup.objects.get(name=auth_group_name)
                        except AuthAttributeGroup.DoesNotExist:
                            errors.append(f"Identifiers Sheet Row {row_num}: Auth attribute group '{auth_group_name}' not found")
                            continue
                    
                    expired_auth_group = None
                    if expired_auth_group_name:
                        try:
                            expired_auth_group = AuthAttributeGroup.objects.get(name=expired_auth_group_name)
                        except AuthAttributeGroup.DoesNotExist:
                            errors.append(f"Identifiers Sheet Row {row_num}: Expired auth attribute group '{expired_auth_group_name}' not found")
                            continue
                    
                    # Parse expiration date
                    expiration_date = None
                    if expiration_date_str:
                        try:
                            expiration_date = datetime.strptime(str(expiration_date_str), '%Y-%m-%d')
                            expiration_date = timezone.make_aware(expiration_date)
                        except ValueError:
                            try:
                                expiration_date = datetime.strptime(str(expiration_date_str), '%Y-%m-%d %H:%M:%S')
                                expiration_date = timezone.make_aware(expiration_date)
                            except ValueError:
                                errors.append(f"Identifiers Sheet Row {row_num}: Invalid expiration date format. Use YYYY-MM-DD or YYYY-MM-DD HH:MM:SS")
                                continue
                    
                    # Create identifier
                    try:
                        identifier_data = {
                            'identifier_type_id': identifier_type.id,
                            'value': value,
                            'plain_password': password or '',
                            'is_enabled': is_enabled,
                            'comment': comment or '',
                            'reject_expired': reject_expired
                        }
                        if auth_group:
                            identifier_data['auth_attribute_group_id'] = auth_group.id
                        if expired_auth_group:
                            identifier_data['expired_auth_attribute_group_id'] = expired_auth_group.id
                        if expiration_date:
                            identifier_data['expiration_date'] = expiration_date
                        
                        serializer = UserIdentifierSerializer(data=identifier_data)
                        if serializer.is_valid():
                            identifier = serializer.save(user=user)
                        else:
                            errors.append(f"Identifiers Sheet Row {row_num}: {serializer.errors}")
                    except Exception as e:
                        errors.append(f"Identifiers Sheet Row {row_num}: Error creating identifier - {str(e)}")
            
            # Process NAS Authorizations sheet
            if "NAS Authorizations" in wb.sheetnames:
                ws_auth = wb["NAS Authorizations"]
                for row_num, row in enumerate(ws_auth.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row):  # Skip empty rows
                        continue
                    
                    user_key = row[0] if len(row) > 0 else None  # Email or External ID
                    identifier_value = row[1] if len(row) > 1 else None
                    nas_name = row[2] if len(row) > 2 else None
                    auth_group_name = row[3] if len(row) > 3 else None
                    
                    if not user_key or not identifier_value or not nas_name:
                        errors.append(f"NAS Authorizations Sheet Row {row_num}: User key, Identifier Value, and NAS Name are required")
                        continue
                    
                    # Find user
                    user = user_map.get(user_key)
                    if not user:
                        try:
                            user = User.objects.get(email=user_key)
                        except User.DoesNotExist:
                            try:
                                user = User.objects.get(external_id=user_key)
                            except User.DoesNotExist:
                                errors.append(f"NAS Authorizations Sheet Row {row_num}: User with key '{user_key}' not found")
                                continue
                    
                    # Find identifier
                    try:
                        identifier = UserIdentifier.objects.get(user=user, value=identifier_value)
                    except UserIdentifier.DoesNotExist:
                        errors.append(f"NAS Authorizations Sheet Row {row_num}: Identifier with value '{identifier_value}' not found for user")
                        continue
                    
                    # Find NAS
                    try:
                        nas = Nas.objects.get(name=nas_name)
                    except Nas.DoesNotExist:
                        errors.append(f"NAS Authorizations Sheet Row {row_num}: NAS with name '{nas_name}' not found")
                        continue
                    
                    # Get auth attribute group
                    auth_group = None
                    if auth_group_name:
                        try:
                            auth_group = AuthAttributeGroup.objects.get(name=auth_group_name)
                        except AuthAttributeGroup.DoesNotExist:
                            errors.append(f"NAS Authorizations Sheet Row {row_num}: Auth attribute group '{auth_group_name}' not found")
                            continue
                    
                    # Create authorization
                    try:
                        auth_data = {
                            'nas': nas.id
                        }
                        if auth_group:
                            auth_data['attribute_group'] = auth_group.id
                        
                        serializer = UserIdentifierNasAuthorizationCreateSerializer(data=auth_data)
                        if serializer.is_valid():
                            serializer.save(user_identifier=identifier)
                        else:
                            errors.append(f"NAS Authorizations Sheet Row {row_num}: {serializer.errors}")
                    except Exception as e:
                        if 'unique' in str(e).lower():
                            # Authorization already exists, skip
                            pass
                        else:
                            errors.append(f"NAS Authorizations Sheet Row {row_num}: Error creating authorization - {str(e)}")
            
            return Response({
                "success": True,
                "created": len(created_users),
                "created_users": created_users,
                "errors": errors
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response(
                {"error": f"Error processing file: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

class UserIdentifierNasAuthorizationViewSet(viewsets.ModelViewSet):
    queryset = UserIdentifierNasAuthorization.objects.all()
    serializer_class = UserIdentifierNasAuthorizationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        identifier_id = self.kwargs.get('identifier_pk')
        return self.queryset.filter(user_identifier_id=identifier_id)

    def get_serializer_class(self):
        if self.action == 'create':
            return UserIdentifierNasAuthorizationCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return UserIdentifierNasAuthorizationUpdateSerializer
        return self.serializer_class

    def perform_create(self, serializer):
        identifier_id = self.kwargs.get('identifier_pk')
        serializer.save(user_identifier_id=identifier_id)

class AvailableNasDevicesView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, identifier_pk):
        # Get all NAS devices that are not authorized for this identifier
        authorized_nas_ids = UserIdentifierNasAuthorization.objects.filter(
            user_identifier_id=identifier_pk
        ).values_list('nas_id', flat=True)
        
        available_nas = Nas.objects.exclude(id__in=authorized_nas_ids)
        serializer = NasSerializer(available_nas, many=True)
        return Response(serializer.data)